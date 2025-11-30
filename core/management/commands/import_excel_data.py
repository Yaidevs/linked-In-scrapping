# core/management/commands/import_excel_data.py
import os
from django.core.management.base import BaseCommand
from core.models import Company, Person, Keyword
import openpyxl

class Command(BaseCommand):
    def handle(self, *args, **options):
        file_path = 'Ejemplodeempresas.xlsx'
        
        if not os.path.exists(file_path):
            self.stdout.write(self.style.ERROR(f'File not found: {file_path}'))
            return
        
        wb = openpyxl.load_workbook(file_path)
        
        # Import from Directores sheet
        sheet = wb['Directores']
        
        for row_num, row in enumerate(sheet.iter_rows(min_row=4, values_only=True), 4):
            # Columns: A=empty, B=Empresa, C=Website, D=Director, E=LinkedIn, F=Memoria
            if len(row) >= 5 and row[1] and row[2]:  # Empresa and Website exist
                company, created = Company.objects.get_or_create(
                    name=row[1],
                    website=row[2]
                )
                self.stdout.write(f"Added company: {row[1]}")
                
                if row[3]:  # Director name exists
                    person, created = Person.objects.get_or_create(
                        name=row[3],
                        company=company,
                        linkedin_url=row[4] if row[4] else ''
                    )
                    self.stdout.write(f"  - Added person: {row[3]}")
        
        # Import Keywords from palabras sheet
        sheet_keywords = wb['palabras']
        
        for row_num, row in enumerate(sheet_keywords.iter_rows(min_row=4, values_only=True), 4):
            # Columns: B, C, D, E contain keywords
            for col_index in [1, 2, 3, 4]:
                if len(row) > col_index and row[col_index]:
                    keyword = str(row[col_index]).strip()
                    if keyword and keyword not in ['Digital transformation', 'None']:
                        Keyword.objects.get_or_create(word=keyword)
                        self.stdout.write(f"Added keyword: {keyword}")
        
        self.stdout.write(self.style.SUCCESS('Excel data imported successfully!'))